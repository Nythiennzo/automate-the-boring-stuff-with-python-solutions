from pathlib import Path
import openpyxl
import csv

working_folder = r'C:\repos\python\automate_the_boring_stuff\chapter16\excelSpreadsheets'
working_folder_path = Path(working_folder)
excel_file_extention = '.xlsx'

for excel_file in working_folder_path.glob('*' + excel_file_extention):
    workbook = openpyxl.load_workbook(excel_file)
    
    csv_file_prefix = excel_file.name.rstrip(excel_file_extention) + '_'

    for sheetName in workbook.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = workbook[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csv_File = open(working_folder_path / (csv_file_prefix + sheetName + '.csv'), 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csv_writer = csv.writer(csv_File)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(rowNum, colNum).value)
            # Write the rowData list to the CSV file.
            csv_writer.writerow(rowData)

        csv_File.close()