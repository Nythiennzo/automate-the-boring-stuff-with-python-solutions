import sys, openpyxl
from openpyxl.styles import Font

source_workbook_filename = sys.argv[1]
try:
    source_workbook = openpyxl.load_workbook(source_workbook_filename)
except:
    print('The first argument was not a valid excel file.')    
    sys.exit()

source_active_sheet = source_workbook.active

after_workbook = openpyxl.load_workbook(source_workbook_filename)
after_active_sheet = after_workbook.active

after_active_sheet.delete_rows(1, after_active_sheet.max_row)
after_active_sheet.delete_cols(1, after_active_sheet.max_column)

for row_number in range(1, source_active_sheet.max_row + 1):
    for column_number in range(1, source_active_sheet.max_column + 1):
        source_cell = source_active_sheet.cell(row_number, column_number)
        target_cell = after_active_sheet.cell(column_number, row_number)

after_workbook.save('inverted_' + source_workbook_filename)
after_workbook.close()