import sys, openpyxl
from openpyxl.styles import Font

try:
    number = int(sys.argv[1])
except:
    print('The first argument was not an integer.')
    sys.exit()


workbook = openpyxl.Workbook()
active_sheet = workbook.active

number_range = range(1, number + 1)
for label_number in number_range:
    bold_font = Font(bold=True)

    row_label_cell = active_sheet.cell(1, label_number + 1)
    row_label_cell.value = label_number
    row_label_cell.font = bold_font  

    column_label_cell = active_sheet.cell(label_number + 1, 1)
    column_label_cell.value = label_number
    column_label_cell.font = bold_font  


for row_number in number_range:
    for column_number in number_range:
        active_sheet.cell(row_number + 1, column_number + 1).value = row_number * column_number

workbook.save('MultiplicationTableFor' + sys.argv[1] + '.xlsx')        
workbook.close()