import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

workbook_path = 'text_files_to_spreadsheet.xlsx'

workbook = openpyxl.load_workbook(workbook_path)
active_sheet = workbook.active

for column_number in range(1, active_sheet.max_column + 1):
    text_file = open('column' + get_column_letter(column_number) + '.txt', 'w')
    
    for row_number in range(1, active_sheet.max_row + 1):
        text_file.write(active_sheet.cell(row_number, column_number).value)

    text_file.close()

workbook.close()